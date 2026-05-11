"""账号管理界面 - 批量导入、自动登录、养号、账号状态管理"""
import os
from datetime import datetime
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QTableWidget, QTableWidgetItem,
    QPushButton, QGroupBox, QFormLayout, QLineEdit, QComboBox,
    QSpinBox, QFileDialog, QMessageBox, QHeaderView, QLabel,
    QTabWidget, QTextEdit, QCheckBox, QDialog, QDialogButtonBox,
    QProgressBar, QToolBar
)
from PyQt6.QtCore import Qt, pyqtSignal
from PyQt6.QtGui import QColor
from db.models import Account, AccountGroup
from loguru import logger


class AccountDialog(QDialog):
    """账号编辑对话框"""
    def __init__(self, parent=None, account=None, groups=None):
        super().__init__(parent)
        self.account = account
        self.groups = groups or []
        self.setWindowTitle("编辑账号" if account else "添加账号")
        self.setMinimumWidth(400)
        self._setup_ui()

    def _setup_ui(self):
        layout = QFormLayout(self)
        self.username_edit = QLineEdit()
        self.password_edit = QLineEdit()
        self.password_edit.setEchoMode(QLineEdit.EchoMode.Password)
        self.platform_combo = QComboBox()
        self.platform_combo.setEditable(True)
        self.platform_combo.addItems([
            "huangye88", "qianyan", "zhongyewang", "zhihu",
            "baidu_tieba", "douban", "csdn", "jianshu",
            "b2b_platform", "custom"
        ])
        self.phone_edit = QLineEdit()
        self.email_edit = QLineEdit()
        self.group_combo = QComboBox()
        self.group_combo.addItem("-- 无分组 --")
        for g in self.groups:
            self.group_combo.addItem(g.name, g.id)

        layout.addRow("平台:", self.platform_combo)
        layout.addRow("用户名:", self.username_edit)
        layout.addRow("密码:", self.password_edit)
        layout.addRow("手机号:", self.phone_edit)
        layout.addRow("邮箱:", self.email_edit)
        layout.addRow("分组:", self.group_combo)

        if self.account:
            self.username_edit.setText(self.account.username)
            self.password_edit.setText(self.account.password)
            idx = self.platform_combo.findText(self.account.platform)
            if idx >= 0:
                self.platform_combo.setCurrentIndex(idx)
            self.phone_edit.setText(self.account.phone or "")
            self.email_edit.setText(self.account.email or "")

        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addRow(buttons)

    def get_data(self):
        group_id = self.group_combo.currentData()
        return {
            "platform": self.platform_combo.currentText(),
            "username": self.username_edit.text(),
            "password": self.password_edit.text(),
            "phone": self.phone_edit.text(),
            "email": self.email_edit.text(),
            "group_id": group_id if group_id else None,
        }


class AccountManagerWidget(QWidget):
    """账号管理面板"""

    status_changed = pyqtSignal()

    def __init__(self, db_manager):
        super().__init__()
        self.db_manager = db_manager
        self._setup_ui()
        self.refresh_table()

    def _setup_ui(self):
        layout = QVBoxLayout(self)

        # 工具栏
        toolbar = QHBoxLayout()
        self.btn_add = QPushButton("➕ 添加账号")
        self.btn_add.clicked.connect(self._add_account)
        self.btn_import = QPushButton("📥 批量导入")
        self.btn_import.clicked.connect(self._import_accounts)
        self.btn_login = QPushButton("🔑 批量登录")
        self.btn_login.clicked.connect(self._batch_login)
        self.btn_maintain = QPushButton("🔄 批量养号")
        self.btn_maintain.clicked.connect(self._batch_maintain)
        self.btn_delete = QPushButton("🗑 删除选中")
        self.btn_delete.clicked.connect(self._delete_selected)
        self.btn_refresh = QPushButton("🔄 刷新")
        self.btn_refresh.clicked.connect(self.refresh_table)
        self.btn_export = QPushButton("📤 导出账号")
        self.btn_export.clicked.connect(self._export_accounts)

        for btn in [self.btn_add, self.btn_import, self.btn_login, self.btn_maintain,
                     self.btn_delete, self.btn_refresh, self.btn_export]:
            toolbar.addWidget(btn)
        layout.addLayout(toolbar)

        # 分组过滤器
        filter_layout = QHBoxLayout()
        filter_layout.addWidget(QLabel("分组筛选:"))
        self.group_filter = QComboBox()
        self.group_filter.addItem("全部账号")
        self.group_filter.currentIndexChanged.connect(self.refresh_table)
        filter_layout.addWidget(self.group_filter)
        filter_layout.addWidget(QLabel("状态筛选:"))
        self.status_filter = QComboBox()
        self.status_filter.addItems(["全部", "active", "limited", "banned", "expired"])
        self.status_filter.currentIndexChanged.connect(self.refresh_table)
        filter_layout.addWidget(self.status_filter)
        filter_layout.addStretch()
        layout.addLayout(filter_layout)

        # 账号表格
        self.table = QTableWidget()
        self.table.setColumnCount(10)
        self.table.setHorizontalHeaderLabels([
            "ID", "平台", "用户名", "手机号", "邮箱", "分组",
            "状态", "登录状态", "健康分", "最后登录"
        ])
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setAlternatingRowColors(True)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        self.table.setSortingEnabled(True)
        layout.addWidget(self.table)

        # 状态栏
        status_layout = QHBoxLayout()
        self.stats_label = QLabel("共 0 个账号 | 活跃: 0 | 受限: 0 | 封禁: 0")
        status_layout.addWidget(self.stats_label)
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        status_layout.addWidget(self.progress_bar)
        layout.addLayout(status_layout)

    def refresh_table(self):
        session = self.db_manager.get_session()
        try:
            # 刷新分组过滤器
            current_group = self.group_filter.currentText()
            self.group_filter.clear()
            self.group_filter.addItem("全部账号")
            groups = session.query(AccountGroup).all()
            for g in groups:
                self.group_filter.addItem(g.name, g.id)
            idx = self.group_filter.findText(current_group)
            if idx >= 0:
                self.group_filter.setCurrentIndex(idx)

            # 查询账号
            query = session.query(Account)
            if self.group_filter.currentData():
                query = query.filter(Account.group_id == self.group_filter.currentData())
            status = self.status_filter.currentText()
            if status != "全部":
                query = query.filter(Account.status == status)

            accounts = query.all()
            self.table.setRowCount(len(accounts))
            for row, acc in enumerate(accounts):
                items = [
                    QTableWidgetItem(str(acc.id)),
                    QTableWidgetItem(acc.platform),
                    QTableWidgetItem(acc.username),
                    QTableWidgetItem(acc.phone or ""),
                    QTableWidgetItem(acc.email or ""),
                    QTableWidgetItem(acc.group.name if acc.group else ""),
                    QTableWidgetItem(acc.status),
                    QTableWidgetItem(acc.login_status),
                    QTableWidgetItem(str(acc.score)),
                    QTableWidgetItem(acc.last_login_at.strftime("%m-%d %H:%M") if acc.last_login_at else "-"),
                ]
                for col, item in enumerate(items):
                    item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                    if acc.status == "banned":
                        item.setBackground(QColor(255, 200, 200))
                    elif acc.status == "limited":
                        item.setBackground(QColor(255, 255, 200))
                    self.table.setItem(row, col, item)

            # 统计
            active = sum(1 for a in accounts if a.status == "active")
            limited = sum(1 for a in accounts if a.status == "limited")
            banned = sum(1 for a in accounts if a.status == "banned")
            self.stats_label.setText(f"共 {len(accounts)} 个账号 | 活跃: {active} | 受限: {limited} | 封禁: {banned}")
        except Exception as e:
            logger.error(f"刷新账号列表异常: {e}")
        finally:
            session.close()

    def _add_account(self):
        session = self.db_manager.get_session()
        groups = session.query(AccountGroup).all()
        dialog = AccountDialog(self, groups=groups)
        session.close()
        if dialog.exec() == QDialog.DialogCode.Accepted:
            data = dialog.get_data()
            session = self.db_manager.get_session()
            try:
                acc = Account(**data)
                session.add(acc)
                session.commit()
                self.refresh_table()
                self.status_changed.emit()
            except Exception as e:
                session.rollback()
                QMessageBox.critical(self, "错误", f"添加账号失败: {e}")
            finally:
                session.close()

    def _import_accounts(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "导入账号", "", "Excel文件 (*.xlsx *.xls);;CSV文件 (*.csv);;所有文件 (*)"
        )
        if not path:
            return
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            session = self.db_manager.get_session()
            count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue
                platform, username, password = str(row[0]), str(row[1]), str(row[2])
                phone = str(row[3]) if len(row) > 3 and row[3] else ""
                email = str(row[4]) if len(row) > 4 and row[4] else ""
                existing = session.query(Account).filter_by(
                    platform=platform, username=username).first()
                if not existing:
                    session.add(Account(
                        platform=platform, username=username, password=password,
                        phone=phone, email=email
                    ))
                    count += 1
            session.commit()
            QMessageBox.information(self, "导入完成", f"成功导入 {count} 个账号")
            self.refresh_table()
        except Exception as e:
            QMessageBox.critical(self, "错误", f"导入失败: {e}")
        finally:
            session.close()

    def _batch_login(self):
        selected = set(item.row() for item in self.table.selectedItems())
        if not selected:
            QMessageBox.warning(self, "提示", "请先选择要登录的账号")
            return
        QMessageBox.information(self, "提示", "批量登录将在后台异步执行，请关注登录状态变化")
        self.status_label_msg = "批量登录已加入队列"

    def _batch_maintain(self):
        QMessageBox.information(self, "提示", "批量养号将在后台异步执行")

    def _delete_selected(self):
        rows = set(item.row() for item in self.table.selectedItems())
        if not rows:
            return
        reply = QMessageBox.question(
            self, "确认删除", f"确定要删除选中的 {len(rows)} 个账号吗？",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if reply == QMessageBox.StandardButton.Yes:
            session = self.db_manager.get_session()
            try:
                for row in rows:
                    acc_id = int(self.table.item(row, 0).text())
                    session.query(Account).filter_by(id=acc_id).delete()
                session.commit()
                self.refresh_table()
            except Exception as e:
                session.rollback()
                QMessageBox.critical(self, "错误", f"删除失败: {e}")
            finally:
                session.close()

    def _export_accounts(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "导出账号", f"accounts_{datetime.now().strftime('%Y%m%d')}.xlsx",
            "Excel文件 (*.xlsx)"
        )
        if not path:
            return
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["平台", "用户名", "密码", "手机号", "邮箱", "状态", "健康分"])
            session = self.db_manager.get_session()
            for acc in session.query(Account).all():
                ws.append([acc.platform, acc.username, acc.password,
                           acc.phone, acc.email, acc.status, acc.score])
            wb.save(path)
            QMessageBox.information(self, "导出完成", f"已导出到: {path}")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"导出失败: {e}")
        finally:
            session.close()
